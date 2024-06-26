from typing import List
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from core.settings import STATIONS
from core.model import Station
from core.dao import (
    today8_waterlevel,
    yesterday8_waterlevel,
    lastweek8_waterlevel,
    lastyear8_waterlevel,
)


# 获取目标水位
async def get_waterlevel() -> List[Station]:
    today8 = await today8_waterlevel()
    yesterday_8 = await yesterday8_waterlevel()
    lastweek8 = await lastweek8_waterlevel()
    lastyear8 = await lastyear8_waterlevel()
    for station in STATIONS:
        # 获取今天 8:00 水位
        for today in today8:
            if station.stcd == today.stcd:
                station.today_8 = today.current
        # 获取昨天 8:00 水位
        for yesterday in yesterday_8:
            if station.stcd == yesterday.stcd:
                station.yesterday_8 = yesterday.current
        # 获取上周 8:00 水位
        for lastweek in lastweek8:
            if station.stcd == lastweek.stcd:
                station.lastweek_8 = lastweek.current
        # 获取去年同期 8:00 水位
        for lastyear in lastyear8:
            if station.stcd == lastyear.stcd:
                station.lastyear_8 = lastyear.current
    return STATIONS


# 保存为xlsx文件
async def save_xlsx(source_file: str, target_file: str, stations: List[Station]):
    wb = load_workbook(source_file)
    ws = wb.active
    ws["A2"] = "填报日期： " + datetime.now().strftime("%Y年%m月%d日")
    for row, station in zip(ws["D5:D14"], stations):
        for cell in row:
            cell.value = station.today_8
            # 达到设防水位 浅蓝色
            if station.today_8 >= station.sfsw and station.today_8 <station.jjsw:
                cell.font = Font(color='189FA7')
            # 达到警戒水位 深蓝色
            elif station.today_8 >= station.jjsw and station.today_8 < station.bzsw:
                cell.font = Font(color='0070C0')
            # 达到保证水位 红色
            elif station.today_8 >= station.bzsw:
                cell.font = Font(color='C00400')
    for row, station in zip(ws["E5:E14"], stations):
        for cell in row:
            cell.value = station.yesterday_8
            if station.yesterday_8 >= station.sfsw and station.yesterday_8 < station.jjsw:
                cell.font = Font(color='189FA7')
            elif station.yesterday_8 >= station.jjsw and station.yesterday_8 < station.bzsw:
                cell.font = Font(color='0070C0')
            elif station.yesterday_8 >= station.bzsw:
                cell.font = Font(color='C00400')
    for row, station in zip(ws["F5:F14"], stations):
        for cell in row:
            cell.value = station.lastweek_8
            if station.lastweek_8 >= station.sfsw and station.lastweek_8 < station.jjsw:
                cell.font = Font(color='189FA7')
            elif station.lastweek_8 >= station.jjsw and station.lastweek_8 < station.bzsw:
                cell.font = Font(color='0070C0')
            elif station.lastweek_8 >= station.bzsw:
                cell.font = Font(color='C00400')
    for row, station in zip(ws["G5:G14"], stations):
        for cell in row:
            cell.value = station.lastyear_8
            if station.lastyear_8 >= station.sfsw and station.lastyear_8 < station.jjsw:
                cell.font = Font(color='189FA7')
            elif station.lastyear_8 >= station.jjsw and station.lastyear_8 < station.bzsw:
                cell.font = Font(color='0070C0')
            elif station.lastyear_8 >= station.bzsw:
                cell.font = Font(color='C00400')
    wb.save(target_file)
