from pathlib import Path
from asyncio import run
from datetime import datetime
from core.service import get_waterlevel, save_csv
from core.settings import STATIONS
from openpyxl import load_workbook


async def main():
    stations = await get_waterlevel(STATIONS)
    filename = "今天、昨天、上周、去年同期8点水位.csv"
    save_csv(filename, stations)


async def main2():
    current_file_path = Path(__file__).resolve()
    current_dir = current_file_path.parent
    file_path = current_dir / "基础表格.xlsx"
    filename = str(file_path)
    print("读取路径: ", filename)
    wb = load_workbook(filename)
    stations = await get_waterlevel(STATIONS)
    file_name: str = datetime.now().strftime("%Y年%m月%d日_鸠江区三线水位测站记录表")
    ws = wb.active
    ws["A2"] = "填报日期： " + datetime.now().strftime("%Y年%m月%d日")
    for row, station in zip(ws["D5:D14"], stations):
        for cell in row:
            cell.value = station.today_8
    for row, station in zip(ws["E5:E14"], stations):
        for cell in row:
            cell.value = station.yesterday_8
    for row, station in zip(ws["F5:F14"], stations):
        for cell in row:
            cell.value = station.lastweek_8
    for row, station in zip(ws["G5:G14"], stations):
        for cell in row:
            cell.value = station.lastyear_8
    savefilename = f"{file_name}.xlsx"
    save_path = current_dir.parent / savefilename
    filename = str(save_path)
    print("保存路径: ", filename)
    wb.save(filename)


if __name__ == "__main__":
    run(main2())
